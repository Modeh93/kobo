"""
KoboDownloader — Streamlit version
Run:  streamlit run kobo_app.py
"""

import streamlit as st
import json, requests, os, time, re, tempfile
from datetime import datetime, timedelta
from collections import Counter, defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    import pyreadstat
    SPSS_OK = True
except ImportError:
    SPSS_OK = False

try:
    import pandas as pd
    PANDAS_OK = True
except ImportError:
    PANDAS_OK = False

CONFIG_FILE = "kobo_downloader_config.json"

# ── File-type buckets ─────────────────────────────────────────────
IMAGE_EXT = {".jpg",".jpeg",".png",".gif",".bmp",".tiff",".tif",".webp",".heic",".heif"}
AUDIO_EXT = {".mp3",".wav",".ogg",".m4a",".aac",".flac",".opus",".amr",".3gp"}
VIDEO_EXT = {".mp4",".mov",".avi",".mkv",".webm",".flv",".wmv"}

def file_subfolder(filename):
    ext = os.path.splitext(filename.lower())[1]
    if ext in IMAGE_EXT: return "images"
    if ext in AUDIO_EXT: return "audio"
    if ext in VIDEO_EXT: return "video"
    return "other"

# ── Config persistence ────────────────────────────────────────────
def load_cfg():
    try:
        with open(CONFIG_FILE) as f: return json.load(f)
    except: return {}

def save_cfg(d):
    try:
        with open(CONFIG_FILE, "w") as f: json.dump(d, f, indent=2)
    except: pass

# ─────────────────────────────────────────────────────────────────
# Backend — all pure functions (no UI dependency)
# ─────────────────────────────────────────────────────────────────
def detect_multi_select_parents(headers):
    counts = Counter()
    for h in headers:
        if not h: continue
        parts = str(h).split("/")
        if len(parts) >= 2:
            counts["/".join(parts[:-1])] += 1
    return {p for p, c in counts.items() if c > 1}

def rename_header(h, mode, multi_parents):
    if not h or mode == "full":
        return str(h) if h else ""
    parts  = str(h).split("/")
    parent = "/".join(parts[:-1])
    if parent in multi_parents and len(parts) >= 3:
        return "/".join(parts[-2:])
    return parts[-1]

def download_kobo_export(token, kobo_url, uid, d_from, d_to, log, prog):
    headers = {
        "Authorization": f"Token {token}",
        "Accept": "application/json",
        "Content-Type": "application/json",
    }
    base    = kobo_url.rstrip("/")
    payload = {
        "type": "xls", "lang": "_default", "group_sep": "/",
        "hierarchy_in_labels": False, "multiple_select": "both",
        "fields_from_all_versions": True,
    }
    if d_from or d_to:
        q = {"_submission_time": {}}
        if d_from: q["_submission_time"]["$gte"] = d_from + "T00:00:00"
        if d_to:   q["_submission_time"]["$lte"] = d_to   + "T23:59:59"
        payload["query"] = q

    log("Creating export job...")
    r = requests.post(f"{base}/api/v2/assets/{uid}/exports/",
                      headers=headers, json=payload, timeout=30)
    if r.status_code != 201:
        raise Exception(r.text)
    export_uid = r.json()["uid"]
    prog(10)

    status_url = f"{base}/api/v2/assets/{uid}/exports/{export_uid}/"
    log("Waiting for Kobo to build the file...")
    for i in range(120):
        time.sleep(2)
        r     = requests.get(status_url, headers=headers)
        r.raise_for_status()
        state = r.json().get("status")
        log(f"  [{i*2}s] {state}")
        prog(10 + min(i, 55))
        if state == "complete":
            file_url = r.json()["result"]
            log("Downloading file...")
            tmp = os.path.join(tempfile.gettempdir(), "kobo_tmp.xlsx")
            with requests.get(file_url,
                              headers={"Authorization": f"Token {token}"},
                              stream=True) as resp:
                resp.raise_for_status()
                with open(tmp, "wb") as f:
                    for chunk in resp.iter_content(8192): f.write(chunk)
            prog(70)
            return tmp
        elif state in ["failed", "error"]:
            raise Exception("Export failed")
    raise Exception("Timeout waiting for export")


def fetch_submissions_with_attachments(token, base, uid, d_from, d_to, log):
    hdrs   = {"Authorization": f"Token {token}", "Accept": "application/json"}
    params = {"format": "json", "limit": 30000}
    if d_from or d_to:
        q = {"_submission_time": {}}
        if d_from: q["_submission_time"]["$gte"] = d_from + "T00:00:00"
        if d_to:   q["_submission_time"]["$lte"] = d_to   + "T23:59:59"
        params["query"] = json.dumps(q)

    log("Fetching submission metadata for attachments...")
    r = requests.get(f"{base}/api/v2/assets/{uid}/data/",
                     headers=hdrs, params=params, timeout=60)
    r.raise_for_status()

    attach_map = {}
    for sub in r.json().get("results", []):
        sub_id   = sub.get("_id")
        sub_uuid = sub.get("_uuid", "")
        atts     = sub.get("_attachments", [])
        if atts:
            for att in atts:
                att["_uuid"] = sub_uuid
            attach_map[sub_id] = atts

    total_files = sum(len(v) for v in attach_map.values())
    log(f"  Found {len(attach_map)} submission(s) with attachments ({total_files} file(s))")
    return attach_map


def download_attachments(token, attach_map, out_dir, log, prog_cb=None):
    base_dir = os.path.join(out_dir, "attachments")
    for sub in ("images", "audio", "video", "other"):
        os.makedirs(os.path.join(base_dir, sub), exist_ok=True)

    hdrs   = {"Authorization": f"Token {token}"}
    result = {}
    total  = sum(len(v) for v in attach_map.values())
    done   = 0
    failed = 0

    log(f"Downloading {total} file(s) → attachments/images|audio|video|other/")

    for sub_id, attachments in attach_map.items():
        result[sub_id] = []
        sub_uuid = attachments[0].get("_uuid", "") if attachments else ""
        prefix   = sub_uuid if sub_uuid else str(sub_id)

        for att in attachments:
            url      = (att.get("download_url") or
                        att.get("download_large_url") or
                        att.get("download_medium_url") or "")
            filename = att.get("filename", "") or att.get("name", "")
            basename = os.path.basename(filename.replace("\\", "/")) or f"file_{done}"
            safe_name  = f"{prefix}_{basename}"
            kind       = file_subfolder(basename)
            local_path = os.path.join(base_dir, kind, safe_name)

            if not url:
                result[sub_id].append({"url": "", "local_path": "", "kind": kind})
                done += 1
                continue

            if os.path.exists(local_path):
                log(f"  ↷ Exists: {kind}/{safe_name}")
                result[sub_id].append({"url": url, "local_path": local_path, "kind": kind})
                done += 1
                if prog_cb: prog_cb(int(done / total * 100) if total else 100)
                continue

            try:
                with requests.get(url, headers=hdrs, stream=True, timeout=60) as resp:
                    resp.raise_for_status()
                    with open(local_path, "wb") as f:
                        for chunk in resp.iter_content(8192): f.write(chunk)
                result[sub_id].append({"url": url, "local_path": local_path, "kind": kind})
            except Exception as e:
                log(f"  ✗ Failed: {safe_name} ({e})")
                result[sub_id].append({"url": url, "local_path": "", "kind": kind})
                failed += 1

            done += 1
            if prog_cb: prog_cb(int(done / total * 100) if total else 100)

    counts  = Counter()
    for atts in result.values():
        for a in atts:
            if a.get("local_path"): counts[a["kind"]] += 1
    if counts:
        log("  ✓ " + ", ".join(f"{counts[k]} {k}" for k in
                                ("images","audio","video","other") if counts[k]))
    if failed:
        log(f"  ✗ {failed} failed")
    return result


def inject_attachments_into_excel(out_path, attach_result, log):
    wb           = openpyxl.load_workbook(out_path)
    ws           = wb.active or wb.worksheets[0]
    bold_font    = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_att = max((len(v) for v in attach_result.values()), default=0)
    if max_att == 0:
        log("  No attachment data to inject.")
        wb.save(out_path); return

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    id_col  = next(
        (i + 1 for i, h in enumerate(headers)
         if str(h or "").strip().lower() in ("_id", "id")), None)

    if id_col is None:
        log("  ⚠ _id column not found — skipping attachment injection.")
        wb.save(out_path); return

    start_col   = ws.max_column + 1
    new_headers = []
    for n in range(1, max_att + 1):
        sfx = f"_{n}" if max_att > 1 else ""
        new_headers += [f"attachment_url{sfx}", f"attachment_local_path{sfx}"]

    for i, h in enumerate(new_headers):
        cell = ws.cell(1, start_col + i, h)
        cell.font = bold_font; cell.alignment = center_align

    for row_idx in range(2, ws.max_row + 1):
        raw = ws.cell(row_idx, id_col).value
        try:    sub_id = int(raw) if raw is not None else None
        except: sub_id = raw
        atts = attach_result.get(sub_id, [])
        col  = start_col
        for n in range(max_att):
            att = atts[n] if n < len(atts) else {}
            ws.cell(row_idx, col,     att.get("url", ""))
            ws.cell(row_idx, col + 1, att.get("local_path", ""))
            col += 2

    for c in range(start_col, start_col + len(new_headers)):
        letter = ws.cell(1, c).column_letter
        mx = max((len(str(ws.cell(r, c).value or ""))
                  for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[letter].width = min(mx + 2, 60)

    wb.save(out_path)
    log(f"  ✓ Added {len(new_headers)} attachment column(s)")


def parse_survey_schema(survey_rows):
    slots        = []
    path_stack   = []
    repeat_stack = []

    def full_path(name):
        return "/".join(path_stack + [name]) if name else ""

    for row in survey_rows:
        t    = row.get("type", "")
        name = row.get("name", "") or row.get("$autoname", "") or ""
        if not t: continue
        if t == "begin_repeat":
            slots.append({"kind": "repeat_begin", "name": name,
                           "full_path": full_path(name),
                           "parent_repeat": repeat_stack[-1] if repeat_stack else None})
            path_stack.append(name); repeat_stack.append(name)
        elif t == "end_repeat":
            if path_stack:   path_stack.pop()
            if repeat_stack: repeat_stack.pop()
            slots.append({"kind": "repeat_end", "name": name})
        elif t == "begin_group":
            path_stack.append(name)
        elif t == "end_group":
            if path_stack: path_stack.pop()
        elif not t.startswith("end_"):
            slots.append({"kind": "field", "name": name,
                           "full_path": full_path(name), "q_type": t,
                           "in_repeat": repeat_stack[-1] if repeat_stack else None})
    return slots


def fetch_survey_schema(token, base, uid):
    hdrs = {"Authorization": f"Token {token}", "Accept": "application/json"}
    r    = requests.get(f"{base}/api/v2/assets/{uid}/?format=json",
                        headers=hdrs, timeout=20)
    r.raise_for_status()
    return parse_survey_schema(r.json().get("content", {}).get("survey", []))


def rebuild_excel(src_path, out_path, fmt, hdr_mode, log, prog, schema=None):
    wb_src      = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    sheet_names = wb_src.sheetnames
    all_sheets  = {}

    for sh in sheet_names:
        ws   = wb_src[sh]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            all_sheets[sh] = {"headers": [], "data": []}; continue
        hdrs = list(rows[0])
        all_sheets[sh] = {"headers": hdrs,
                           "data": [dict(zip(hdrs, r)) for r in rows[1:]]}
        log(f"  Read: {sh} ({len(all_sheets[sh]['data'])} rows, {len(hdrs)} cols)")
    wb_src.close()

    main_name    = sheet_names[0]
    repeat_names = sheet_names[1:]
    main_data    = all_sheets[main_name]["data"]
    main_headers = all_sheets[main_name]["headers"]
    prog(75)

    if fmt == "wide":
        meta_cols = {"_index","_parent_index","_parent_table_name","_submission_time",
                     "_uuid","_id","_status","_submitted_by","_notes","_tags",
                     "_validation_status","_geolocation","_attachments"}

        all_hdrs_flat = list(main_headers)
        for rn in repeat_names: all_hdrs_flat += all_sheets[rn]["headers"]
        multi_parents = detect_multi_select_parents(all_hdrs_flat)

        repeat_info = {}
        for rname in repeat_names:
            rdata    = all_sheets[rname]["data"]
            r_fields = [h for h in all_sheets[rname]["headers"] if h and h not in meta_cols]
            groups   = defaultdict(list)
            for row in rdata: groups[row.get("_parent_index")].append(row)
            mx = max((len(v) for v in groups.values()), default=0)
            repeat_info[rname] = {"fields": r_fields, "groups": groups, "max": mx}
            log(f"  {rname}: {mx} max repetitions")

        final_cols = []

        if schema:
            def find_sheet(rn):
                for sh in repeat_names:
                    if rn.lower() in sh.lower(): return sh
                return None

            def add_repeat_fields(sname):
                if not sname or sname not in repeat_info: return
                ri = repeat_info[sname]
                for n in range(1, ri["max"] + 1):
                    for fh in ri["fields"]:
                        base = rename_header(fh, hdr_mode, multi_parents)
                        final_cols.append(("repeat", fh, f"{base}_{n}", sname, n))

            added_repeats   = set()
            main_hdr_lookup = {str(h): h for h in main_headers if h}

            for slot in schema:
                if slot["kind"] == "field" and slot["in_repeat"] is None:
                    fp, name = slot["full_path"], slot["name"]
                    matched  = main_hdr_lookup.get(fp) or next(
                        (h for h in main_headers if h and (
                            str(h) == fp or str(h).endswith("/" + fp)
                            or str(h).split("/")[-1] == name)), None)
                    if matched and matched not in {c[1] for c in final_cols}:
                        final_cols.append(("main", matched,
                                           rename_header(matched, hdr_mode, multi_parents)))
                elif slot["kind"] == "repeat_begin":
                    sh = find_sheet(slot["name"])
                    if sh and sh not in added_repeats:
                        added_repeats.add(sh); add_repeat_fields(sh)

            covered = {c[1] for c in final_cols if c[0] == "main"}
            for h in main_headers:
                if h and h not in covered:
                    final_cols.append(("main", h, rename_header(h, hdr_mode, multi_parents)))
        else:
            repeat_parent = {}
            msn = sheet_names[0]
            for rname in repeat_names:
                rdata = all_sheets[rname]["data"]
                repeat_parent[rname] = rdata[0].get("_parent_table_name", msn) if rdata else msn
            for rname in repeat_names:
                ia     = len(main_headers) - 1
                parent = repeat_parent[rname]
                ri_f   = repeat_info[rname]["fields"]
                if parent == msn and ri_f:
                    parts = str(ri_f[0]).split("/")
                    for depth in range(len(parts), 0, -1):
                        pfx = "/".join(parts[:depth])
                        for i, h in enumerate(main_headers):
                            if h and str(h).startswith(pfx): ia = i
                        if ia != len(main_headers) - 1: break
                repeat_info[rname]["insert_after"]  = ia
                repeat_info[rname]["nested_parent"] = (
                    parent if parent != msn and parent in repeat_info else None)

            insert_map = defaultdict(list)
            for rname in sorted(
                [r for r in repeat_names if repeat_info[r]["nested_parent"] is None],
                key=lambda r: repeat_info[r]["insert_after"]):
                insert_map[repeat_info[rname]["insert_after"]].append(rname)

            def add_repeat_cols(rname):
                ri = repeat_info[rname]
                for n in range(1, ri["max"] + 1):
                    for fh in ri["fields"]:
                        base = rename_header(fh, hdr_mode, multi_parents)
                        final_cols.append(("repeat", fh, f"{base}_{n}", rname, n))
                for child in repeat_names:
                    if repeat_info[child]["nested_parent"] == rname:
                        add_repeat_cols(child)

            for i, h in enumerate(main_headers):
                if h:
                    final_cols.append(("main", h, rename_header(h, hdr_mode, multi_parents)))
                for rname in insert_map.get(i, []): add_repeat_cols(rname)

        out_rows = []
        for rec in main_data:
            pidx = rec.get("_index") or -1
            row  = {}
            for col in final_cols:
                if col[0] == "main":
                    _, sk, disp = col
                    row[disp] = rec.get(sk, "")
                else:
                    _, sk, disp, rname, n = col
                    reps = repeat_info[rname]["groups"].get(pidx, [])
                    rep  = reps[n-1] if n-1 < len(reps) else {}
                    row[disp] = rep.get(sk, "")
            out_rows.append(row)

        out_sheet_list = [{"name": "Data", "data": out_rows,
                           "headers": [c[2] for c in final_cols]}]
    else:
        out_sheet_list = []
        for sh in sheet_names:
            hdrs      = all_sheets[sh]["headers"]
            mp        = detect_multi_select_parents(hdrs)
            disp_hdrs = [rename_header(h, hdr_mode, mp) if h else "" for h in hdrs]
            out_sheet_list.append({"name": sh[:31], "data": all_sheets[sh]["data"],
                                   "headers": hdrs, "display": disp_hdrs})

    prog(88)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    bold_font    = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for sd in out_sheet_list:
        ws   = wb_out.create_sheet(title=sd["name"])
        rows = sd["data"]
        hdrs = sd.get("headers", [])
        disp = sd.get("display", hdrs)
        if not rows:
            ws.append(["No data"]); continue
        for ci, h in enumerate(disp, 1):
            cell = ws.cell(1, ci, h)
            cell.font = bold_font; cell.alignment = center_align
        ws.row_dimensions[1].height = 22
        for ri, rec in enumerate(rows, 2):
            for ci, key in enumerate(hdrs, 1):
                v = rec.get(key, "") if isinstance(rec, dict) else ""
                if v is None: v = ""
                elif isinstance(v, (dict, list)): v = json.dumps(v, ensure_ascii=False)
                ws.cell(ri, ci, v)
        for col_obj in ws.columns:
            mx = max((len(str(c.value or "")) for c in col_obj), default=8)
            ws.column_dimensions[col_obj[0].column_letter].width = min(mx+2, 40)
        ws.freeze_panes = "A2"

    wb_out.save(out_path)
    log(f"✓ Saved: {out_path}")


def rebuild_to_sheets(src_path, fmt, hdr_mode, log, prog, schema=None):
    tmp_out = os.path.join(tempfile.gettempdir(), "kobo_rebuilt.xlsx")
    try:
        rebuild_excel(src_path, tmp_out, fmt, hdr_mode, log, prog, schema=schema)
        wb2    = openpyxl.load_workbook(tmp_out, read_only=True, data_only=True)
        result = []
        for sh in wb2.sheetnames:
            ws    = wb2[sh]
            rows2 = list(ws.iter_rows(values_only=True))
            if not rows2:
                result.append({"name": sh, "rows": [], "headers": []}); continue
            hdrs2 = [str(h) if h else "" for h in rows2[0]]
            result.append({"name": sh, "rows": [dict(zip(hdrs2, r)) for r in rows2[1:]],
                            "headers": hdrs2})
        wb2.close()
        return result
    finally:
        if os.path.exists(tmp_out): os.remove(tmp_out)


def write_spss(sheets_data, out_path):
    import pandas as pd, pyreadstat

    def safe_varname(name, seen):
        clean = re.sub(r'[^a-zA-Z0-9_]', '_', str(name))
        if not clean or not clean[0].isalpha(): clean = "v_" + clean
        clean = clean[:64]
        orig, i = clean, 1
        while clean in seen:
            sfx = f"_{i}"; clean = orig[:64-len(sfx)] + sfx; i += 1
        seen.add(clean)
        return clean

    base    = out_path[:-4] if out_path.endswith(".sav") else out_path
    written = 0
    for si, sd in enumerate(sheets_data):
        rows, hdrs, sh_name = sd["rows"], sd["headers"], sd["name"]
        if not rows: continue
        sav_path   = out_path if si == 0 else f"{base}_{sh_name}.sav"
        seen       = set()
        safe_names = [safe_varname(h, seen) for h in hdrs]
        records    = []
        for rec in rows:
            row = []
            for h in hdrs:
                v = rec.get(h, "") if isinstance(rec, dict) else ""
                if v is None: v = ""
                elif isinstance(v, (dict, list)): v = json.dumps(v, ensure_ascii=False)
                row.append(str(v) if v != "" else "")
            records.append(row)
        df         = pd.DataFrame(records, columns=safe_names)
        var_labels = {s: o for s, o in zip(safe_names, hdrs)}
        saved = False
        for kw in ["variable_labels", "column_labels", "variable_display_labels"]:
            try:
                pyreadstat.write_sav(df, sav_path, **{kw: var_labels})
                saved = True; break
            except TypeError: continue
        if not saved: pyreadstat.write_sav(df, sav_path)
        written += 1
    return written


# ─────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="KoboDownloader",
        page_icon="📥",
        layout="centered",
    )

    st.title("📥 KoboDownloader")
    st.caption("Kobo → Excel / SPSS  |  Long / Wide  |  Attachments")
    st.divider()

    # ── Load saved config ────────────────────────────────────────
    cfg = load_cfg()

    # ── KOBO SERVER ──────────────────────────────────────────────
    st.subheader("🌐 Kobo Server")
    kobo_url = st.selectbox(
        "Server URL",
        options=[
            "https://kobo.unhcr.org",
            "https://kf.kobotoolbox.org",
            "https://kobo.humanitarianresponse.info",
        ],
        index=["https://kobo.unhcr.org",
               "https://kf.kobotoolbox.org",
               "https://kobo.humanitarianresponse.info"].index(
            cfg.get("url", "https://kobo.unhcr.org"))
        if cfg.get("url") in ["https://kobo.unhcr.org",
                               "https://kf.kobotoolbox.org",
                               "https://kobo.humanitarianresponse.info"] else 0,
        label_visibility="collapsed",
    )

    # ── API TOKEN ────────────────────────────────────────────────
    st.subheader("🔑 API Token")
    token = st.text_input(
        "Token", value=cfg.get("token", ""),
        type="password", label_visibility="collapsed",
        placeholder="Paste your Kobo API token here...",
    )

    # ── LOAD FORMS ───────────────────────────────────────────────
    if st.button("📋 Load Forms", use_container_width=True):
        if not token:
            st.error("Enter API token first.")
        else:
            with st.spinner("Loading forms..."):
                try:
                    r = requests.get(
                        kobo_url.rstrip("/") + "/api/v2/assets/?asset_type=survey",
                        headers={"Authorization": f"Token {token}"}, timeout=20)
                    r.raise_for_status()
                    forms = r.json()["results"]
                    st.session_state["forms"] = {
                        f"{p['name']}  ({p.get('deployment__submission_count',0)} submissions)": p
                        for p in forms
                    }
                    st.success(f"✓ {len(forms)} form(s) loaded")
                except Exception as e:
                    st.error(f"Failed: {e}")

    # ── FORM SELECT ──────────────────────────────────────────────
    st.subheader("📄 Form")
    forms_dict = st.session_state.get("forms", {})
    form_label = st.selectbox(
        "Select form", options=list(forms_dict.keys()),
        label_visibility="collapsed",
        placeholder="Load forms first...",
        disabled=not forms_dict,
    )
    project = forms_dict.get(form_label)
    if project:
        st.caption(f"UID: `{project['uid']}`  —  "
                   f"{project.get('deployment__submission_count','?')} submissions")

    # ── DATE RANGE ───────────────────────────────────────────────
    st.subheader("📅 Date Range  (optional)")
    col1, col2, col3, col4, col5 = st.columns([1,1,1,1,2])
    if col1.button("Today"):
        st.session_state["d_from"] = datetime.now().strftime("%Y-%m-%d")
        st.session_state["d_to"]   = datetime.now().strftime("%Y-%m-%d")
    if col2.button("7 days"):
        st.session_state["d_from"] = (datetime.now()-timedelta(days=7)).strftime("%Y-%m-%d")
        st.session_state["d_to"]   = datetime.now().strftime("%Y-%m-%d")
    if col3.button("30 days"):
        st.session_state["d_from"] = (datetime.now()-timedelta(days=30)).strftime("%Y-%m-%d")
        st.session_state["d_to"]   = datetime.now().strftime("%Y-%m-%d")
    if col4.button("All"):
        st.session_state["d_from"] = ""
        st.session_state["d_to"]   = ""

    dcol1, dcol2 = st.columns(2)
    d_from = dcol1.text_input("From (YYYY-MM-DD)",
                               value=st.session_state.get("d_from", ""),
                               placeholder="2024-01-01")
    d_to   = dcol2.text_input("To   (YYYY-MM-DD)",
                               value=st.session_state.get("d_to", ""),
                               placeholder="2024-12-31")

    # ── FORMAT OPTIONS ───────────────────────────────────────────
    st.subheader("⚙️ Format Options")
    fcol1, fcol2 = st.columns(2)

    with fcol1:
        st.markdown("**Structure**")
        fmt = st.radio(
            "Structure", label_visibility="collapsed",
            options=["long", "wide"],
            format_func=lambda x: "Long — تاب لكل repeat group" if x == "long"
                                  else "Wide — كل شيء في تاب واحدة",
            index=0 if cfg.get("fmt","long") == "long" else 1,
        )

    with fcol2:
        st.markdown("**Column Names**")
        hdr = st.radio(
            "Column names", label_visibility="collapsed",
            options=["short", "full"],
            format_func=lambda x: "Short — بدون اسم الجروب" if x == "short"
                                  else "Full — مع اسم الجروب كاملاً",
            index=0 if cfg.get("hdr","short") == "short" else 1,
        )

    # ── OUTPUT FORMAT ────────────────────────────────────────────
    st.subheader("💾 Output Format")
    out_fmt = st.radio(
        "Output format", label_visibility="collapsed",
        options=["excel", "spss"],
        format_func=lambda x: "Excel (.xlsx)" if x == "excel" else "SPSS (.sav)",
        horizontal=True,
    )

    # ── OUTPUT PATH ──────────────────────────────────────────────
    st.subheader("📁 Output Path  (on server)")
    default_out = cfg.get("out", os.path.join(os.path.expanduser("~"),
                                               "kobo_export.xlsx"))
    if out_fmt == "spss":
        default_out = default_out.replace(".xlsx", ".sav")
        if not default_out.endswith(".sav"):
            default_out = os.path.splitext(default_out)[0] + ".sav"
    else:
        default_out = default_out.replace(".sav", ".xlsx")
        if not default_out.endswith(".xlsx"):
            default_out = os.path.splitext(default_out)[0] + ".xlsx"

    out_path = st.text_input(
        "Output path", value=default_out,
        label_visibility="collapsed",
        help="Full path on the server where the file will be saved",
    )

    # ── ATTACHMENTS ──────────────────────────────────────────────
    st.subheader("📎 Attachments")
    dl_att = st.checkbox(
        "Download attachments",
        value=cfg.get("att", False),
        help="Downloads files into attachments/images|audio|video|other/ next to the output file",
    )
    if dl_att:
        att_dir = os.path.join(os.path.dirname(os.path.abspath(out_path)), "attachments")
        st.info(
            f"📁 `{att_dir}/`\n"
            "```\n"
            "├── 🖼  images/\n"
            "├── 🔊  audio/\n"
            "├── 🎬  video/\n"
            "└── 📄  other/\n"
            "```\n"
            "Columns added to Excel: `attachment_url` · `attachment_local_path`"
        )
        if out_fmt == "spss":
            st.warning("⚠️ Attachments are only injected into Excel files, not SPSS.")

    st.divider()

    # ── DOWNLOAD BUTTON ──────────────────────────────────────────
    if not EXCEL_OK:
        st.error("Missing: `pip install openpyxl`")
    if out_fmt == "spss" and not SPSS_OK:
        st.error("Missing: `pip install pyreadstat pandas`")

    ready = bool(token and project)
    if st.button("⬇️  Download", use_container_width=True,
                 type="primary", disabled=not ready):

        save_cfg({"url": kobo_url, "token": token, "out": out_path,
                  "fmt": fmt, "hdr": hdr, "att": dl_att})

        # ── Live log + progress ──────────────────────────────────
        log_box  = st.empty()
        prog_bar = st.progress(0)
        log_lines = []

        def log(msg):
            ts = datetime.now().strftime("%H:%M:%S")
            log_lines.append(f"[{ts}] {msg}")
            log_box.code("\n".join(log_lines[-30:]))  # show last 30 lines

        def prog(v):
            prog_bar.progress(min(int(v), 100))

        uid = project["uid"]

        try:
            log(f"Form: {project['name']}")
            log(f"Format: {fmt.upper()}  |  Headers: {hdr.capitalize()}"
                + ("  |  +Attachments" if dl_att else ""))
            log(f"Date: {d_from or 'all'} → {d_to or 'all'}")

            # 1. Download raw export
            tmp = download_kobo_export(token, kobo_url, uid, d_from, d_to, log, prog)

            # 2. Fetch schema
            schema = None
            try:
                log("Fetching survey schema...")
                schema = fetch_survey_schema(token, kobo_url.rstrip("/"), uid)
                log(f"  Schema: {len(schema)} elements")
            except Exception as se:
                log(f"  Schema unavailable ({se}) — using fallback ordering")

            # 3. Process & write output
            log(f"Processing ({fmt} / {out_fmt})...")
            os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)

            if out_fmt == "spss":
                sheets = rebuild_to_sheets(tmp, fmt, "short", log, prog, schema=schema)
                os.remove(tmp)
                log("Writing SPSS .sav file(s)...")
                n = write_spss(sheets, out_path)
                if n and n > 1:
                    log(f"  Note: {n} files saved")
            else:
                rebuild_excel(tmp, out_path, fmt, hdr, log, prog, schema=schema)
                os.remove(tmp)

            # 4. Attachments
            if dl_att and out_fmt == "excel":
                log("─── Attachments ───────────────────")
                attach_map = fetch_submissions_with_attachments(
                    token, kobo_url.rstrip("/"), uid, d_from, d_to, log)
                if attach_map:
                    out_dir = os.path.dirname(os.path.abspath(out_path))
                    attach_result = download_attachments(
                        token, attach_map, out_dir, log,
                        prog_cb=lambda v: prog(v))
                    log("Adding attachment columns to Excel...")
                    inject_attachments_into_excel(out_path, attach_result, log)
                else:
                    log("  No attachments found.")

            prog(100)
            log(f"✅ Done! Saved to: {out_path}")

            # ── Download button for browser ──────────────────────
            if out_fmt == "excel" and os.path.exists(out_path):
                with open(out_path, "rb") as f:
                    st.download_button(
                        label="⬇️  Download Excel file",
                        data=f,
                        file_name=os.path.basename(out_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
            elif out_fmt == "spss" and os.path.exists(out_path):
                with open(out_path, "rb") as f:
                    st.download_button(
                        label="⬇️  Download SPSS file",
                        data=f,
                        file_name=os.path.basename(out_path),
                        mime="application/octet-stream",
                        use_container_width=True,
                    )
            st.success(f"✅ File saved on server: `{out_path}`")

        except Exception as e:
            log(f"✗ Error: {e}")
            st.error(f"Failed: {e}")


if __name__ == "__main__":
    main()
